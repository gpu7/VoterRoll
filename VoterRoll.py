# Author:  Richard Casey
# Date:    10-05-2024 (DD-MM-YYYY)
# Purpose: VoterRoll compares county voter roll files with the National Change of Address (NCOA) database 
#          and identifies voters who moved out-of-state or out-of-country.

# Standard library imports
import os
import shutil
import sys
import warnings
from   typing import List

# Third-party imports
import pandas as pd
from   openpyxl                 import load_workbook
from   openpyxl.styles          import PatternFill
from   openpyxl.utils.dataframe import dataframe_to_rows
from   pandas                   import DataFrame

# Local application/library specific imports
from utilities.loggerUtilVoterRoll import logger

# suppress the warning "Workbook contains no default style, apply openpyxl's default"
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# suppress specific FutureWarnings related to pandas DataFrame concatenation
warnings.filterwarnings('ignore', category=FutureWarning, message='The behavior of DataFrame concatenation with empty or all-NA entries is deprecated.')

# directories and files
BASE_DIR                  = os.path.dirname(os.path.abspath(__file__))
COLORADO_VOTERS_MOVED_DIR = os.path.join(BASE_DIR, "colorado_voters_moved")
VOTERS_MOVED_FILE         = os.path.join(BASE_DIR, "voters_moved.xlsx")

# list of Colorado counties
colorado_counties: List[str] = [
    "Adams_5",       "Alamosa_31",    "Arapahoe_3",  "Archuleta_34", "Baca_56",         "Bent_50",
    "Boulder_9",     "Broomfield_12", "Chaffee_26",  "Cheyenne_59",   "Clear Creek_39",
    "Conejos_42",    "Costilla_55",   "Crowley_47",  "Custer_52",     "Delta_19",       "Denver_2",
    "Dolores_58",    "Douglas_6",     "Eagle_15",    "El Paso_1",     "Elbert_21",      "Fremont_16",
    "Garfield_13",   "Gilpin_48",     "Grand_32",    "Gunnison_30",   "Hinsdale_63",    "Huerfano_43",
    "Jackson_61",    "Jefferson_4",   "Kiowa_60",    "Kit Carson_44", "Lake_41",        "LaPlata_14",
    "Larimer_7",     "Las Animas_33", "Lincoln_49",  "Logan_25",      "Mesa_11",        "Mineral_62",
    "Moffat_35",     "Montezuma_22",  "Montrose_17", "Morgan_20",     "Otero_27",       "Ouray_51",
    "Park_28",       "Phillips_54",   "Pitkin_29",   "Prowers_36",    "Pueblo_10",      "Rio Blanco_45",
    "Rio Grande_37", "Routt_23",      "Saguache_46", "San Juan_64",   "San Miguel_40",
    "Sedgwick_57",   "Summit_18",     "Teller_24",   "Washington_53", "Weld_8",         "Yuma_38"
]

# freeze first row in Excel file
def freeze_first_row(file_path):
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        ws.freeze_panes = ws['A2']
        wb.save(file_path)
    except Exception as e:
        logger.error(f"ERROR: failed to freeze first row in {file_path}: {e}")

# autofit columns in Excel file
def autofit_columns(file_path):
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 5)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
    except Exception as e:
        logger.error(f"ERROR: failed to autofit columns in {file_path}: {e}")

# color rows by voters who moved out-of-state or out-of-country
def color_rows(file_path):
    try:
        wb = load_workbook(file_path)
        ws = wb.active

        # fill styles for shades of grey - from light to dark
        gainsboro_grey_fill  = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")
        silver_grey_fill     = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        dark_grey_fill       = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")

        # Get the column indexes for MAILING_STATE and MAILING_COUNTRY
        mailing_state_col   = None
        mailing_country_col = None
        for cell in ws[1]:  # Iterate over the first row to find the columns
            if cell.value   == "MAILING_STATE":
                mailing_state_col   = cell.column
            elif cell.value == "MAILING_COUNTRY":
                mailing_country_col = cell.column

        if not mailing_state_col or not mailing_country_col:
            logger.error("ERROR: MAILING_STATE or MAILING_COUNTRY column not found.")
            sys.exit(1)

        # Color rows based on MAILING_STATE and MAILING_COUNTRY
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            mailing_state_value   = row[mailing_state_col   - 1].value
            mailing_country_value = row[mailing_country_col - 1].value

            if mailing_state_value:                                     # color rows for voters who moved out-of-state
                for cell in row:
                    cell.fill = gainsboro_grey_fill
            elif mailing_country_value:                                 # color rows for voters who moved out-of-country
                for cell in row:
                    cell.fill = silver_grey_fill
            elif not mailing_state_value and not mailing_country_value: # color rows where there are no values in MAILING_STATE or MAILING_COUNTRY
                for cell in row:
                    cell.fill = dark_grey_fill

        wb.save(file_path)
    except Exception as e:
        logger.error(f"ERROR: failed to color rows in {file_path}: {e}")

# main
def main() -> None:
    # check if voters_moved.xlsx exists in same directory as VoterRoll.py
    if not os.path.exists(VOTERS_MOVED_FILE):
        logger.error("ERROR: 'voters_moved.xlsx' file does not exist.")
        sys.exit(1)
    
    # Copy voters_moved.xlsx to each county directory
    logger.info("Copy voters_moved.xlsx to Colorado county directories...")
    try:
        for sub_dir in colorado_counties:
            sub_dir_path: str = os.path.join(BASE_DIR, sub_dir)
            if not os.path.isdir(sub_dir_path):
                logger.warning(f"WARNING: {sub_dir_path} is not a valid directory.")
                continue
            dest_path: str = os.path.join(BASE_DIR, sub_dir, "voters_moved.xlsx")
            logger.debug(f"Copy voters_moved.xlsx to {dest_path}")
            pd.read_excel(VOTERS_MOVED_FILE).to_excel(dest_path, index=False)
    except Exception as e:
        logger.error(f"ERROR: error copying voters_moved.xlsx file: {e}")
        sys.exit(1)
    
    # Rename voters_moved.xlsx to county_name_voters_moved.xlsx
    logger.info("Rename voters_moved.xlsx to county_name_voters_moved.xlsx...")
    try:
        for sub_dir in colorado_counties:
            sub_dir_path: str = os.path.join(BASE_DIR, sub_dir)
            if not os.path.isdir(sub_dir_path):
                logger.warning(f"WARNING: {sub_dir_path} is not a valid directory.")
                continue
            old_path: str = os.path.join(sub_dir_path, "voters_moved.xlsx")
            new_path: str = os.path.join(sub_dir_path, f"{sub_dir}_voters_moved.xlsx")
            logger.debug(f"Rename {old_path} to {new_path}")
            if os.path.exists(old_path):
                if os.path.exists(new_path):
                    os.remove(new_path)
                os.rename(old_path, new_path)
            else:
                logger.warning(f"WARNING: file {old_path} does not exist.")
    except Exception as e:
        logger.error(f"ERROR: error renaming files: {e}")
        sys.exit(1)

    # process each county directory
    logger.info("Process Colorado county directories...")
    try:
        for sub_dir in colorado_counties:
            sub_dir_path: str = os.path.join(BASE_DIR, sub_dir)
            if not os.path.isdir(sub_dir_path):
                logger.warning(f"WARNING: {sub_dir_path} is not a valid directory.")
                continue
            
            # Add _voters_moved.xlsx suffix to county file
            county_file: str = os.path.join(sub_dir_path, f"{sub_dir}_voters_moved.xlsx")
            if not os.path.exists(county_file):
                logger.warning(f"File {county_file} does not exist.")
                continue
            county_df: DataFrame = pd.read_excel(county_file)
            logger.info(f"Processing: {os.path.basename(county_file)}")

            # search for NCOA file in county directory
            ncoa_files: List[str] = [f for f in os.listdir(sub_dir_path) if 'NCOA' in f and f.endswith('.xlsx')]
            if not ncoa_files:
                logger.info(f"No NCOA file found in {sub_dir}.")
                continue

            # Sort the NCOA files by date and select the most recent one
            ncoa_files.sort(reverse=True, key=lambda x: x[:8])  # Sort by the date part of the filename
            ncoa_file: str = ncoa_files[0]
            ncoa_df: DataFrame = pd.read_excel(os.path.join(sub_dir_path, ncoa_file))
            logger.info(f"Processing NCOA file: {ncoa_file} in {sub_dir}")
            
            # search for voter roll (VR) file in county directory
            vr_files: List[str] = [f for f in os.listdir(sub_dir_path) if f.startswith('VR') and f.endswith('.xlsx')]
            if not vr_files:
                logger.info(f"No voter roll (VR) file found in {sub_dir}.")
                continue

            # Sort the VR files by date and select the most recent one
            vr_files.sort(reverse=True, key=lambda x: x[2:9])  # Sort by the date part of the filename
            vr_file: str = vr_files[0]
            vr_df: DataFrame = pd.read_excel(os.path.join(sub_dir_path, vr_file))
            logger.info(f"Processing voter roll (VR) file: {vr_file} in {sub_dir}")

            # search for voters who moved out-of-state or out-of-country
            for _, row in ncoa_df.iterrows():
                if row['NEW State'] != 'CO':
                    voter_id: int = row['VoterID']
                    matching_record: DataFrame = vr_df[vr_df['VOTER_ID'] == voter_id]
                    county_df = pd.concat([county_df, matching_record], ignore_index=True)
            
            # sort by MAILING_STATE, then by MAILING_COUNTRY
            county_df = county_df.sort_values(by=["MAILING_STATE", "MAILING_COUNTRY"])

            # format date columns to YYYY-MM-DD without time
            county_df['EFFECTIVE_DATE']         = pd.to_datetime(county_df['EFFECTIVE_DATE']).dt.strftime('%m/%d/%Y')
            county_df['REGISTRATION_DATE']      = pd.to_datetime(county_df['REGISTRATION_DATE']).dt.strftime('%m/%d/%Y')
            county_df['PARTY_AFFILIATION_DATE'] = pd.to_datetime(county_df['PARTY_AFFILIATION_DATE']).dt.strftime('%m/%d/%Y')

            # save the county file
            county_df.to_excel(county_file, index=False)

            # copy county file from each county directory into colorado_voters_moved directory
            if not os.path.exists(COLORADO_VOTERS_MOVED_DIR): # Check if COLORADO_VOTERS_MOVED_DIR exists. If not, create it.
                os.makedirs(COLORADO_VOTERS_MOVED_DIR)
            moved_county_file: str = os.path.join(COLORADO_VOTERS_MOVED_DIR, f"{sub_dir}_voters_moved.xlsx")
            shutil.copy2(county_file, moved_county_file)
            logger.info(f"Copy {moved_county_file} to colorado_voters_moved directory...")

            # format Excel file
            logger.info(f"Format {moved_county_file}...")
            try:
                # TODO: 
                # open and save file to ensure it's in correct format
                # apparently, there are incompatibilities between Excel and LibreOffice Calc
                # need to check if these two lines are necessary
                df = pd.read_excel(moved_county_file)
                df.to_excel(moved_county_file, index=False)

                freeze_first_row(moved_county_file)
                autofit_columns (moved_county_file)
                color_rows      (moved_county_file)
            except Exception as e:
                logger.error(f"ERROR: failed to format {moved_county_file}: {e}")

    except Exception as e:
        logger.error(f"ERROR: error during processing of county directories: {e}")
        sys.exit(1)

# main entry point
if __name__ == "__main__":
    logger.info("Starting VoterRoll...")
    main()
    logger.info("Finished VoterRoll.")